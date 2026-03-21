import streamlit as st
import pandas as pd
from streamlit_gsheets import GSheetsConnection
import math
import html
import time
import io
import json
from datetime import datetime, timedelta
import streamlit.components.v1 as components

# 💡 한국 표준시(KST) 기준 시간 반환
def get_kst_now():
    return datetime.utcnow() + timedelta(hours=9)

# 1. 페이지 기본 설정 (와이드 모드 유지)
st.set_page_config(page_title="유니매입가격정보 - 인상공문 검색", page_icon="📈", layout="wide")

# ==========================================
# 🎨 TOmBOy's INOUT 스타일 다크 테마 커스텀 CSS
# ==========================================
st.markdown("""
    <style>
    /* 스트림릿 기본 UI 요소를 완벽하게 숨기기 */
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header {visibility: hidden !important;}
    [data-testid="stHeader"] {display: none !important;}
    
    /* 전체 여백 최적화 및 다크 테마 배경 */
    [data-testid="stAppViewContainer"] { background-color: #2b323c; }
    .main .block-container { padding-top: 1rem; max-width: 98%; }
    
    /* 텍스트 기본 색상 흰색 (가독성 최적화) */
    h1, h2, h3, h4, p, span, label { color: #ffffff !important; }
    
    /* 버튼 공통 스타일 */
    div.stButton > button {
        border-radius: 8px !important;
        font-weight: bold !important;
        padding: 0px 10px !important;
    }
    
    /* Primary 버튼 (검색, 새로고침 등) -> 파란색 */
    button[kind="primary"] {
        background-color: #4e8cff !important;
        border-color: #4e8cff !important;
        color: white !important;
    }
    button[kind="primary"]:hover { 
        background-color: #3b76e5 !important; 
        border-color: #3b76e5 !important; 
    }

    /* Secondary 버튼 (초기화, 로그아웃 등) -> 올리브색 */
    button[kind="secondary"] {
        background-color: #757c43 !important;
        border-color: #757c43 !important;
        color: white !important;
    }
    button[kind="secondary"]:hover { 
        background-color: #646a39 !important; 
        border-color: #646a39 !important; 
    }
    
    /* 💡 타이틀 버튼 (Tertiary) 스타일링 - 제목 클릭 시 전체보기용 */
    button[kind="tertiary"] {
        display: flex !important;
        justify-content: flex-start !important;
        padding: 0 !important;
        background-color: transparent !important;
        border: none !important;
        margin-top: 5px !important;
    }
    button[kind="tertiary"] p {
        font-size: 1.8rem !important; /* H3 타이틀 크기 */
        font-weight: 800 !important;
        color: #ffffff !important;
        margin: 0 !important;
        padding: 0 !important;
    }
    button[kind="tertiary"]:hover p {
        color: #4e8cff !important; /* 마우스 오버 시 파란색으로 변함 */
    }
    
    /* 💡 EXCEL 다운로드 버튼 전용 올리브색 커스텀 */
    div[data-testid="stDownloadButton"] button[kind="primary"] {
        background-color: #757c43 !important;
        border-color: #757c43 !important;
        color: white !important;
    }
    div[data-testid="stDownloadButton"] button[kind="primary"]:hover {
        background-color: #646a39 !important;
        border-color: #646a39 !important;
        color: white !important;
    }
    
    /* 검색 메뉴 굵게 및 색상 (검색창은 밝게 유지하여 가독성 확보) */
    div[data-testid="stTextInput"] input {
        color: #1e293b !important;
        font-weight: bold !important;
        background-color: #f8f9fa !important;
    }
    div[data-baseweb="select"] > div { 
        font-weight: bold !important; 
        background-color: #f8f9fa !important;
        color: #1e293b !important;
    }
    div[data-baseweb="select"] span { color: #1e293b !important; }
    
    /* 구분선 라인 색상 */
    hr { margin: 15px 0px 15px 0px; border: 0.5px solid #4a5568 !important; }
    </style>
""", unsafe_allow_html=True)

# ==========================================
# 🔒 로그인 기능 (비밀번호 확인 + 3시간 세션 유지)
# ==========================================
def check_password():
    TIMEOUT_SECONDS = 10800  # 3시간 유지
    
    if st.query_params.get("auth") == "true":
        try:
            login_ts = float(st.query_params.get("ts", 0))
            if time.time() - login_ts < TIMEOUT_SECONDS:
                st.session_state["authenticated"] = True
            else:
                st.session_state["authenticated"] = False
                st.query_params.clear()
        except:
            st.session_state["authenticated"] = False
            st.query_params.clear()

    if "authenticated" not in st.session_state:
        st.session_state["authenticated"] = False

    if not st.session_state["authenticated"]:
        st.markdown("<h1 style='text-align: center; color: #4e8cff !important; margin-top: 50px;'>🛡️ ADMIN ACCESS</h1>", unsafe_allow_html=True)
        col_l, col_c, col_r = st.columns([1, 1.2, 1])
        with col_c:
            st.info("유니매입가격정보를 열람하시려면 비밀번호를 입력해주세요.")
            with st.form("login_form"):
                pwd = st.text_input("PASSWORD", type="password", placeholder="••••")
                submit = st.form_submit_button("SYSTEM LOGIN", use_container_width=True, type="primary")
                
                if submit:
                    if pwd == str(st.secrets.get("tom_password", "")):
                        st.session_state["authenticated"] = True
                        st.query_params["auth"] = "true"
                        st.query_params["ts"] = str(time.time())
                        st.rerun()
                    else:
                        st.error("🚨 비밀번호가 일치하지 않습니다.")
        return False
    return True

if not check_password():
    st.stop()

# 2. 구글 시트 데이터 불러오기 (실시간 5초 캐시 유지)
@st.cache_data(ttl=5)
def load_data():
    conn = st.connection("gsheets", type=GSheetsConnection)
    df = conn.read(worksheet="시트1", ttl=5) 
    
    # 에러 방지: 완전 빈 행 제거 및 이름 없는 열 제거
    df = df.dropna(how="all")
    df = df.loc[:, ~df.columns.str.contains('^Unnamed')] 
    df.columns = df.columns.astype(str).str.strip()
    return df

try:
    data = load_data()
except Exception as e:
    st.error("구글 시트를 불러오는 데 실패했습니다. '.streamlit/secrets.toml' 설정과 공유 상태를 확인해주세요.")
    st.stop()

col_vendor = "업체명"
col_item = "물품명"
col_date = "인상날짜"

required_cols = [col_vendor, col_item, col_date]
missing_cols = [col for col in required_cols if col not in data.columns]

if missing_cols:
    st.error(f"🚨 컬럼명 오류: 구글 시트에서 '{', '.join(missing_cols)}' 열을 찾을 수 없습니다.")
    st.info(f"💡 현재 시트에 있는 실제 1열 제목들: {', '.join(map(str, data.columns))}")
    st.stop()

data = data.fillna("")

# ==========================================
# 🛠️ 상태 관리 (검색/전체보기 전용 로직)
# ==========================================
# 1) 실제로 화면 필터링에 반영될 저장된 검색 조건들 (초기값: 전체)
if 'active_vendor' not in st.session_state: st.session_state.active_vendor = "전체"
if 'active_item' not in st.session_state: st.session_state.active_item = "전체"
if 'active_year' not in st.session_state: st.session_state.active_year = "전체"

# 2) UI 껍데기(드롭다운/입력창) 상태
if 'ui_vendor' not in st.session_state: st.session_state.ui_vendor = "전체"
if 'ui_item' not in st.session_state: st.session_state.ui_item = "전체"
if 'ui_year' not in st.session_state: st.session_state.ui_year = "전체"

def do_search():
    """검색 버튼을 눌렀을 때만 UI값을 실제 필터값으로 복사 적용 후 입력창 비우기"""
    st.session_state.active_vendor = st.session_state.ui_vendor
    st.session_state.active_item = st.session_state.ui_item
    st.session_state.active_year = st.session_state.ui_year
    
    # 💡 조건은 저장했으니 드롭다운 선택창은 다시 '전체'로 깔끔하게 초기화합니다.
    st.session_state.ui_vendor = "전체"
    st.session_state.ui_item = "전체"
    st.session_state.ui_year = "전체"

def do_reset():
    """모든 조건과 입력창을 싹 다 비우는 완전 초기화(전체보기) 함수"""
    st.session_state.active_vendor = "전체"
    st.session_state.active_item = "전체"
    st.session_state.active_year = "전체"
    st.session_state.ui_vendor = "전체"
    st.session_state.ui_item = "전체"
    st.session_state.ui_year = "전체"

def do_full_refresh():
    """타이틀 클릭 시: 최신 데이터 갱신 + 완전 초기화"""
    load_data.clear() # 캐시 강제 삭제 (최신 데이터 불러오기)
    do_reset()        # 검색 조건도 모두 비우기


# ==========================================
# UI 레이아웃 시작 (상단 상태바)
# ==========================================
col_t, col_l = st.columns([8.5, 1.5])
with col_t: 
    # 클릭 가능한 타이틀 (클릭 시 최신 데이터 갱신 + 초기 화면 로직 실행)
    st.button("📈 유니매입가격정보 (인상공문 현황)", type="tertiary", help="클릭하면 최신 데이터를 불러오고 전체보기 화면으로 돌아갑니다.", on_click=do_full_refresh)

with col_l:
    if st.button("🔓 LOGOUT", use_container_width=True, type="secondary"):
        st.session_state["authenticated"] = False
        st.query_params.clear()
        st.rerun()

st.markdown("<hr>", unsafe_allow_html=True)

# ==========================================
# 드롭다운 고유 항목 리스트 추출
# ==========================================
years_set = set()
if col_date in data.columns:
    for d in data[col_date].dropna().unique():
        s = str(d).strip()
        if s and s.lower() != 'nan':
            parts = s.replace('-', '.').replace('/', '.').split('.')
            if parts and parts[0].isdigit():
                y = parts[0]
                if len(y) == 2: years_set.add("20" + y)  # '26' -> '2026'
                elif len(y) == 4: years_set.add(y)       # '2026' -> '2026'
year_list = ["전체"] + [f"{y}년" for y in sorted(list(years_set), reverse=True)]

vendor_raw = [str(v).strip() for v in data[col_vendor].unique() if str(v).strip() != ""]
vendor_list = ["전체"] + sorted(vendor_raw)

item_raw = [str(v).strip() for v in data[col_item].unique() if str(v).strip() != ""]
item_list = ["전체"] + sorted(item_raw)


# 1. 상세 검색 영역 (5단 배치: 업체명 / 물품명 / 연도 / 검색 / 전체보기)
search_col1, search_col2, search_col3, search_col4, search_col5 = st.columns([2.5, 2.5, 1.5, 1.2, 1.2])

with search_col1:
    # Selectbox를 사용하여 드롭다운 리스트 및 텍스트 검색 동시 지원
    st.selectbox("🏢 업체명 검색", vendor_list, key="ui_vendor")

with search_col2:
    st.selectbox("📦 물품명 검색", item_list, key="ui_item")

with search_col3:
    st.selectbox("📅 인상 연도", year_list, key="ui_year")

with search_col4:
    st.markdown("<div style='margin-top: 28px;'></div>", unsafe_allow_html=True)
    # 검색 버튼을 누를 때만 동작함 (엔터키 자동 검색 방지)
    st.button("🔍 검색", use_container_width=True, type="primary", on_click=do_search)

with search_col5:
    st.markdown("<div style='margin-top: 28px;'></div>", unsafe_allow_html=True)
    st.button("📂 전체보기", use_container_width=True, type="secondary", on_click=do_reset)

# 2. 필터링 로직 (AND 조건)
filtered_df = data.copy()

# 업체명 필터 적용 ("전체"가 아닐 때만)
if st.session_state.active_vendor != "전체":
    filtered_df = filtered_df[filtered_df[col_vendor].astype(str).str.contains(st.session_state.active_vendor, case=False, na=False)]

# 물품명 필터 적용 ("전체"가 아닐 때만, 위 결과에 이어서 교집합(AND) 적용)
if st.session_state.active_item != "전체":
    filtered_df = filtered_df[filtered_df[col_item].astype(str).str.contains(st.session_state.active_item, case=False, na=False)]

# 연도 필터 적용
if st.session_state.active_year != "전체":
    target_y = st.session_state.active_year.replace("년", "") # "2026"
    target_y_short = target_y[2:]           # "26"
    
    def is_match_year(d):
        s = str(d).strip()
        if not s or s.lower() == 'nan': return False
        parts = s.replace('-', '.').replace('/', '.').split('.')
        return parts[0] == target_y or parts[0] == target_y_short

    filtered_df = filtered_df[filtered_df[col_date].apply(is_match_year)]

# 초기 정렬 설정 (파이썬)
sort_cols = [c for c in [col_date, col_vendor, col_item] if c in filtered_df.columns]
if sort_cols:
    asc_rules = [False if c == col_date else True for c in sort_cols]
    filtered_df = filtered_df.sort_values(by=sort_cols, ascending=asc_rules)


# 3. 요약 지표 (Metrics) 및 [인쇄/엑셀] 가로 일체형 레이아웃
latest_date = "-"
if not filtered_df.empty and col_date in filtered_df.columns:
    valid_dates = [d for d in filtered_df[col_date].tolist() if str(d).strip() != ""]
    if valid_dates:
        latest_date = max(valid_dates)

valid_vendors_cnt = 0
if not filtered_df.empty:
    valid_vendors = [v for v in filtered_df[col_vendor].unique() if str(v).strip() != ""]
    valid_vendors_cnt = len(valid_vendors)

st.markdown("<br>", unsafe_allow_html=True)

col_bar, col_print, col_excel = st.columns([7, 1.5, 1.5])
with col_bar:
    st.markdown(f"""
        <div style="background-color: #353b48; height: 42px; padding: 0 20px; border-radius: 8px; color: #ffffff; font-size: 15px; display: flex; justify-content: space-around; align-items: center; border: 1px solid #4a5568;">
            <span>🔍 총 검색 건수 : <span style="color: #4e8cff; font-size: 18px; font-weight: bold;">{len(filtered_df):,}</span> 건</span>
            <span style="color: #4a5568;">|</span>
            <span>📅 최근 인상 : <span style="color: #4e8cff; font-size: 18px; font-weight: bold;">{latest_date}</span></span>
            <span style="color: #4a5568;">|</span>
            <span>🏢 업체 수 : <span style="color: #4e8cff; font-size: 18px; font-weight: bold;">{valid_vendors_cnt:,}</span> 개사</span>
        </div>
    """, unsafe_allow_html=True)


# --- 🖨️ PRINT 기능용 HTML 생성 ---
html_table = filtered_df.to_html(index=False, escape=True)
html_table = html_table.replace('border="1" class="dataframe"', 'class="custom-table"')

# 프린트용 화면에서도 칸 너비 강제 조정 적용 (물품명 18%, 메모 34%, 기존가날짜 8%)
html_table = html_table.replace('<th>물품명</th>', '<th style="width: 18%;">물품명</th>')
html_table = html_table.replace('<th>메모</th>', '<th style="width: 34%;">메모</th>')
html_table = html_table.replace('<th>기존가날짜</th>', '<th style="width: 8%;">기존가날짜</th>')

print_html_content = f"""
<!DOCTYPE html>
<html>
<head>
    <title>인쇄 미리보기</title>
    <meta charset="utf-8">
    <style>
        @page {{ size: A4 portrait; margin: 10mm; }}
        body {{ font-family: 'Malgun Gothic', 'Apple SD Gothic Neo', sans-serif; color: black; background: white; margin: 0; }}
        .title {{ font-size: 22px; font-weight: bold; text-align: center; margin-bottom: 20px; }}
        .info {{ font-size: 12px; margin-bottom: 10px; color: #555; text-align: right; }}
        .custom-table {{ width: 100%; border-collapse: collapse; font-size: 11.5px; text-align: center; }}
        .custom-table th, .custom-table td {{ border: 1px solid #aaa; padding: 6px 8px; color: black !important; }}
        .custom-table th {{ font-weight: bold; background-color: #f1f5f9; -webkit-print-color-adjust: exact; print-color-adjust: exact; }}
    </style>
</head>
<body>
    <div class="title">유니매입가격정보 검색 결과</div>
    <div class="info">출력 일자: {get_kst_now().strftime('%Y-%m-%d %H:%M')} | 총 검색 건수: {len(filtered_df):,}건</div>
    {html_table}
</body>
</html>
"""

with col_print:
    components.html(
        f"""
        <!DOCTYPE html>
        <html><head>
        <style>
            body {{ margin: 0; padding: 0; overflow: hidden; background-color: transparent; }}
            .btn-print {{
                width: 100%; height: 42px; background-color: #757c43; color: white;
                border: 1px solid #757c43; border-radius: 8px; font-weight: bold; cursor: pointer; font-size: 15px;
                font-family: 'Malgun Gothic', 'Apple SD Gothic Neo', sans-serif;
                display: flex; align-items: center; justify-content: center; box-sizing: border-box;
                margin-top: 0px;
            }}
            .btn-print:hover {{ background-color: #646a39; border-color: #646a39; }}
        </style>
        <script>
            function fastPrint() {{
                const htmlContent = {json.dumps(print_html_content)};
                let iframe = document.getElementById('print-frame');
                if (iframe) {{ document.body.removeChild(iframe); }}
                
                iframe = document.createElement('iframe');
                iframe.id = 'print-frame';
                iframe.style.position = 'absolute'; iframe.style.width = '1px'; iframe.style.height = '1px'; 
                iframe.style.opacity = '0'; iframe.style.pointerEvents = 'none';
                document.body.appendChild(iframe);
                
                const doc = iframe.contentWindow.document;
                doc.open(); doc.write(htmlContent); doc.close();
                
                setTimeout(function() {{ 
                    iframe.contentWindow.focus(); 
                    iframe.contentWindow.print(); 
                }}, 150);
            }}
        </script>
        </head>
        <body>
            <button class="btn-print" onclick="fastPrint()">🖨️ PRINT</button>
        </body></html>
        """, height=45
    )

with col_excel:
    # --- 💾 EXCEL 추출 로직 ---
    excel_output = io.BytesIO()
    has_openpyxl = False
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        has_openpyxl = True
    except ImportError:
        pass
        
    if has_openpyxl and not filtered_df.empty:
        try:
            with pd.ExcelWriter(excel_output, engine='openpyxl') as writer:
                filtered_df.to_excel(writer, index=False, sheet_name='검색결과')
                ws = writer.sheets['검색결과']
                
                font_white = Font(color="FFFFFF", bold=True)
                align_center = Alignment(horizontal="center", vertical="center")
                border_thin = Border(left=Side(style='thin', color='D0D0D0'), right=Side(style='thin', color='D0D0D0'), 
                                     top=Side(style='thin', color='D0D0D0'), bottom=Side(style='thin', color='D0D0D0'))
                
                for col_idx, col_name in enumerate(filtered_df.columns, 1):
                    c_lower = str(col_name).lower()
                    if "기존" in c_lower: fill_color = "3B5B88"
                    elif "인상" in c_lower: fill_color = "B8860B"
                    elif "메모" in c_lower: fill_color = "757C43"
                    else: fill_color = "353B48"
                    
                    cell = ws.cell(row=1, column=col_idx)
                    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                    cell.font = font_white
                    cell.alignment = align_center
                    cell.border = border_thin
                    
                    col_letter = openpyxl.utils.get_column_letter(col_idx)
                    # 엑셀 셀 너비도 뷰 비율에 맞춰 조정 (물품명 확대, 메모 축소, 기존가날짜 확대)
                    if "물품명" in c_lower: ws.column_dimensions[col_letter].width = 18
                    elif "메모" in c_lower: ws.column_dimensions[col_letter].width = 34
                    elif "기존가날짜" in c_lower: ws.column_dimensions[col_letter].width = 11
                    else: ws.column_dimensions[col_letter].width = 15
                    
                for row_idx in range(2, len(filtered_df) + 2):
                    for col_idx in range(1, len(filtered_df.columns) + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.border = border_thin
                        cell.alignment = Alignment(vertical="center")

            excel_data = excel_output.getvalue()
            st.download_button("💾 EXCEL", data=excel_data, file_name=f"유니매입단가인상_{get_kst_now().strftime('%Y%m%d')}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True, type="primary")
        except Exception as e:
            st.error(f"엑셀 생성 오류: {e}")
    else:
        if not filtered_df.empty:
            csv_data = filtered_df.to_csv(index=False).encode('utf-8-sig')
            st.download_button("💾 EXCEL (기본)", data=csv_data, file_name=f"유니매입단가인상_{get_kst_now().strftime('%Y%m%d')}.csv", mime="text/csv", use_container_width=True, type="primary")
        else:
            st.download_button("💾 EXCEL", data="", disabled=True, use_container_width=True, type="primary")


# ==========================================
# 5. 데이터프레임 헤더 (타이틀 및 검색 조건 표시)
# ==========================================
cond_texts = []
if st.session_state.active_vendor != "전체": cond_texts.append(f"업체({st.session_state.active_vendor})")
if st.session_state.active_item != "전체": cond_texts.append(f"물품({st.session_state.active_item})")
if st.session_state.active_year != "전체": cond_texts.append(f"연도({st.session_state.active_year})")

if cond_texts:
    search_info = f"<span style='font-size: 15px; color: #ffeb3b; margin-left: 10px;'>[검색조건: {' + '.join(cond_texts)}]</span>"
else:
    search_info = f"<span style='font-size: 14px; color: #cbd5e1; font-weight: normal; margin-left: 10px;'>(전체 데이터)</span>"

st.markdown(f"<h4 style='color: #ffffff; margin-bottom: 5px; margin-top: 15px;'>📋 상세 내역 {search_info}</h4>", unsafe_allow_html=True)


# ==========================================
# 6. 독립 HTML 테이블 렌더링
# ==========================================
if filtered_df.empty:
    st.warning("👀 검색 조건에 맞는 데이터가 없습니다. 다른 조건으로 검색해 보세요.")
else:
    def get_th_class(col_name):
        c_lower = str(col_name)
        if "기존" in c_lower: return "th-in"     
        if "인상" in c_lower: return "th-out"    
        if "메모" in c_lower: return "th-etc"    
        return "th-base"                         

    def get_td_class(col_name):
        c_lower = str(col_name)
        if "업체" in c_lower or "물품" in c_lower or "메모" in c_lower: return "tl"
        if "가" in c_lower or "폭" in c_lower or "수량" in c_lower: return "tr"
        return "tc"

    bold_cols = ["물품명", "인상폭"]

    iframe_html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="utf-8">
        <style>
            body {{
                margin: 0; padding: 0; 
                font-family: 'Malgun Gothic', 'Apple SD Gothic Neo', sans-serif;
                background-color: #2b323c; 
            }}
            .custom-table-container {{ width: 100%; margin-top: 5px; }}
            .custom-table {{ width: 100%; border-collapse: collapse; font-size: 15px; background-color: white; table-layout: auto; }}
            .custom-table th, .custom-table td {{ border: 1px solid #d0d0d0; padding: 8px 10px; color: #1e293b; }}
            
            .custom-table th {{ 
                text-align: center; color: white !important; font-weight: bold; 
                padding: 10px 6px; cursor: pointer; user-select: none; transition: 0.2s;
            }}
            
            .custom-table tr:nth-child(even) td {{ background-color: #f8f9fa; }}
            .custom-table tr:hover td {{ background-color: #e2e6ea; }}
            .th-base {{ background-color: #353b48 !important; }}
            .th-in {{ background-color: #3b5b88 !important; }}
            .th-out {{ background-color: #b8860b !important; }}
            .th-etc {{ background-color: #757c43 !important; }}
            
            .tc {{ text-align: center; }} .tl {{ text-align: left; }} .tr {{ text-align: right; }}
            .bold-col {{ font-weight: 900 !important; color: #000000 !important; }}
            .sort-icon {{ color: #ffeb3b; margin-left: 5px; font-size: 12px; }}
            
            .page-btn {{
                padding: 6px 16px; font-size: 15px; font-weight: bold; cursor: pointer;
                border: 1px solid #4e8cff; border-radius: 6px; background-color: #4e8cff; color: white; transition: 0.2s;
            }}
            .page-btn:hover:not(:disabled) {{ background-color: #3b76e5; border-color: #3b76e5; }}
            .page-btn:disabled {{ opacity: 0.4; cursor: not-allowed; background-color: #4a5568; border-color: #4a5568; }}
        </style>
    </head>
    <body>
        <div class="custom-table-container">
            <table class='custom-table' id='myTable'>
                <thead><tr>
    """
    
    for i, col in enumerate(filtered_df.columns):
        th_class = get_th_class(col)
        
        # 💡 물품명 칸 크기 확대, 메모 칸 크기 소폭 축소, 기존가날짜 유지 확대 비율
        width_style = ""
        if "물품명" in str(col):
            width_style = "width: 18%;"
        elif "메모" in str(col):
            width_style = "width: 34%;"
        elif "기존가날짜" in str(col):
            width_style = "width: 8%;"
            
        iframe_html += f"<th class='{th_class}' style='{width_style}' onclick='sortTable({i})' title='클릭하여 정렬'>{html.escape(str(col))} <span class='sort-icon' id='icon-{i}'></span></th>"
    iframe_html += "</tr></thead><tbody id='tableBody'>"
    
    # --- 🖥️ 화면 렌더링 속도 최적화 (itertuples 사용) ---
    rows_html = []
    cols_list = filtered_df.columns.tolist()
    
    for idx, row in enumerate(filtered_df.itertuples(index=False)):
        display_style = "" if idx < 100 else " style='display: none;'"
        row_str = f"<tr{display_style}>"
        
        for col_idx, col_name in enumerate(cols_list):
            val = row[col_idx]
            safe_val = "" if pd.isna(val) or val == "" else html.escape(str(val))
            
            td_class = get_td_class(col_name)
            if col_name in bold_cols:
                td_class += " bold-col"
                
            row_str += f"<td class='{td_class}'>{safe_val}</td>"
        row_str += "</tr>"
        rows_html.append(row_str)
        
    iframe_html += "".join(rows_html)
        
    iframe_html += """
                </tbody>
            </table>
        </div>
        
        <!-- 하단 페이지네이션 컨트롤 -->
        <div id='paginationControls' style='display: none; text-align: center; margin-top: 25px; margin-bottom: 25px;'>
            <button class='page-btn' onclick='changePage(-1)' id='prevBtn'>◀ 이전</button>
            <span id='pageInfo' style='margin: 0 25px; font-weight: bold; font-size: 16px; color: #ffffff;'></span>
            <button class='page-btn' onclick='changePage(1)' id='nextBtn'>다음 ▶</button>
        </div>

        <script>
            const ROWS_PER_PAGE = 100;
            let currentPage = 1;
            let sortCol = -1;
            let sortAsc = true;
            let allRows = [];

            window.onload = function() {
                const tbody = document.getElementById("tableBody");
                allRows = Array.from(tbody.getElementsByTagName("tr"));
                updateTable();
            };

            function updateTable() {
                const tbody = document.getElementById("tableBody");
                const totalPages = Math.ceil(allRows.length / ROWS_PER_PAGE) || 1;
                
                if (currentPage < 1) currentPage = 1;
                if (currentPage > totalPages) currentPage = totalPages;

                tbody.innerHTML = "";
                const startIdx = (currentPage - 1) * ROWS_PER_PAGE;
                const endIdx = startIdx + ROWS_PER_PAGE;
                
                for(let i = startIdx; i < endIdx && i < allRows.length; i++) {
                    allRows[i].style.display = ""; 
                    tbody.appendChild(allRows[i]);
                }

                document.getElementById("pageInfo").innerText = currentPage + " / " + totalPages;
                document.getElementById("prevBtn").disabled = (currentPage === 1);
                document.getElementById("nextBtn").disabled = (currentPage === totalPages);
                document.getElementById("paginationControls").style.display = (allRows.length > ROWS_PER_PAGE) ? "block" : "none";
            }

            function changePage(delta) {
                currentPage += delta;
                updateTable();
                window.scrollTo(0, 0); 
            }

            function sortTable(n) {
                if (sortCol === n) { sortAsc = !sortAsc; } 
                else { sortCol = n; sortAsc = true; }
                
                document.querySelectorAll('.sort-icon').forEach(icon => icon.innerHTML = '');
                const iconEl = document.getElementById('icon-' + n);
                if (iconEl) { iconEl.innerHTML = sortAsc ? "▲" : "▼"; }
                
                allRows.sort((a, b) => {
                    const tdA = a.getElementsByTagName("td")[n];
                    const tdB = b.getElementsByTagName("td")[n];
                    const valA = tdA ? tdA.innerText.trim() : "";
                    const valB = tdB ? tdB.innerText.trim() : "";
                    
                    if (valA === "" && valB !== "") return 1;
                    if (valB === "" && valA !== "") return -1;
                    
                    const cleanA = valA.replace(/,/g, '').replace(/원/g, '').replace(/%/g, '').replace(/ /g, '');
                    const cleanB = valB.replace(/,/g, '').replace(/원/g, '').replace(/%/g, '').replace(/ /g, '');
                    
                    const numA = parseFloat(cleanA);
                    const numB = parseFloat(cleanB);
                    
                    let cmpA = (!isNaN(numA) && cleanA === numA.toString()) ? numA : valA;
                    let cmpB = (!isNaN(numB) && cleanB === numB.toString()) ? numB : valB;
                    
                    if (cmpA < cmpB) return sortAsc ? -1 : 1;
                    if (cmpA > cmpB) return sortAsc ? 1 : -1;
                    return 0;
                });
                
                currentPage = 1;
                updateTable();
            }
        </script>
    </body>
    </html>
    """

    # 컴포넌트의 높이를 데이터 개수에 맞춰 동적으로 계산하여 스크롤바 방지
    num_rows = min(len(filtered_df), 100)
    iframe_height = 80 + (num_rows * 39) + (100 if len(filtered_df) > 100 else 30)
    
    # 스트림릿 전용 컴포넌트로 안전하고 완벽하게 HTML/JS 표 렌더링
    components.html(iframe_html, height=iframe_height, scrolling=False)
